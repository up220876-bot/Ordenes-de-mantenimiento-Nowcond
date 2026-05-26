from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import Flask, render_template, request, redirect, url_for, send_file, session
from flask_sqlalchemy import SQLAlchemy


app = Flask(__name__)
app.secret_key = "clave_mantenimiento_correctivo"

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///mantenimiento.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


# =========================
# MODELOS DE BASE DE DATOS
# =========================

class Solicitud(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    folio = db.Column(db.String(20))

    area = db.Column(db.String(100))
    maquina = db.Column(db.String(100))
    fecha_solicitud = db.Column(db.String(50))
    hora_solicitud = db.Column(db.String(50))
    solicitante = db.Column(db.String(100))
    tecnico = db.Column(db.String(100))

    tiempo_ciclo = db.Column(db.String(50))
    piezas_ciclo = db.Column(db.String(50))
    modelo = db.Column(db.String(100))
    minutos_afectacion = db.Column(db.String(50))
    piezas_afectadas = db.Column(db.String(50))

    paro_produccion = db.Column(db.String(50))
    fecha_paro = db.Column(db.String(50))
    hora_paro = db.Column(db.String(50))
    fecha_arranque = db.Column(db.String(50))
    hora_arranque = db.Column(db.String(50))

    descripcion = db.Column(db.Text)

    participo_1 = db.Column(db.String(100))
    participo_2 = db.Column(db.String(100))
    participo_3 = db.Column(db.String(100))

    inicio_actividades = db.Column(db.String(50))
    fin_actividades = db.Column(db.String(50))
    servicio_hrs = db.Column(db.String(50))
    servicio_hh = db.Column(db.String(50))
    tipo_servicio = db.Column(db.String(50))
    tipo_reparacion = db.Column(db.String(50))

    causa_falla = db.Column(db.Text)
    actividad_realizada = db.Column(db.Text)
    refacciones_utilizadas = db.Column(db.Text)
    observaciones_tecnico = db.Column(db.Text)

    maquina_entregada = db.Column(db.String(20))
    fecha_entrega = db.Column(db.String(50))
    observaciones_supervisor = db.Column(db.Text)

    estado = db.Column(db.String(50), default="Activa")


class Area(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), unique=True)


class Supervisor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), unique=True)


class Tecnico(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), unique=True)


class CampoFormato(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre_campo = db.Column(db.String(100), unique=True)
    etiqueta = db.Column(db.String(200))
    seccion = db.Column(db.String(100))
    visible = db.Column(db.Boolean, default=True)
    obligatorio = db.Column(db.Boolean, default=False)
    tipo_campo = db.Column(db.String(50), default="texto")
    es_extra = db.Column(db.Boolean, default=False)


class ValorCampoExtra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    solicitud_id = db.Column(db.Integer, db.ForeignKey("solicitud.id"))
    campo_id = db.Column(db.Integer, db.ForeignKey("campo_formato.id"))
    valor = db.Column(db.Text)


class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100))
    rol = db.Column(db.String(50))
    password = db.Column(db.String(100))
    primer_login = db.Column(db.Boolean, default=True)


# =========================
# FUNCIONES AUXILIARES
# =========================

def generar_folio():
    ultimo = Solicitud.query.order_by(Solicitud.id.desc()).first()
    numero = ultimo.id + 1 if ultimo else 1
    return f"SMC-{numero:04}"


def crear_datos_iniciales():
    if Area.query.count() == 0:
        db.session.add_all([
            Area(nombre="Diecasting"),
            Area(nombre="Sandcasting"),
            Area(nombre="Lavado"),
            Area(nombre="Acabados"),
            Area(nombre="Almacén"),
            Area(nombre="Moldes"),
            Area(nombre="Planta tratadora")
        ])

    if Supervisor.query.count() == 0:
        db.session.add_all([
            Supervisor(nombre="Supervisor 1"),
            Supervisor(nombre="Supervisor 2"),
            Supervisor(nombre="Supervisor 3")
        ])

    if Tecnico.query.count() == 0:
        db.session.add_all([
            Tecnico(nombre="Técnico 1"),
            Tecnico(nombre="Técnico 2"),
            Tecnico(nombre="Técnico 3")
        ])

    if Usuario.query.count() == 0:
        db.session.add(Usuario(
            nombre="Administrador",
            rol="admin",
            password="1234",
            primer_login=True
        ))

    db.session.commit()


def crear_campos_formato_iniciales():
    campos_base = [
        ("area", "Área", "Supervisor"),
        ("maquina", "Máquina o equipo", "Supervisor"),
        ("fecha_solicitud", "Fecha de solicitud", "Supervisor"),
        ("hora_solicitud", "Hora de solicitud", "Supervisor"),
        ("solicitante", "Solicitante", "Supervisor"),
        ("tecnico", "Técnico asignado", "Supervisor"),
        ("tiempo_ciclo", "Tiempo ciclo", "Supervisor"),
        ("piezas_ciclo", "Piezas / ciclo", "Supervisor"),
        ("modelo", "Modelo", "Supervisor"),
        ("minutos_afectacion", "Minutos de afectación", "Supervisor"),
        ("piezas_afectadas", "Piezas afectadas", "Supervisor"),
        ("paro_produccion", "Paro de producción", "Supervisor"),
        ("fecha_paro", "Fecha de paro", "Supervisor"),
        ("hora_paro", "Hora de paro", "Supervisor"),
        ("fecha_arranque", "Fecha de arranque", "Supervisor"),
        ("hora_arranque", "Hora de arranque", "Supervisor"),
        ("descripcion", "Descripción de la falla", "Supervisor"),
        ("foto", "Foto de evidencia", "Supervisor"),

        ("participo_1", "Participó técnico 1", "Técnico"),
        ("participo_2", "Participó técnico 2", "Técnico"),
        ("participo_3", "Participó técnico 3", "Técnico"),
        ("inicio_actividades", "Inicio de actividades", "Técnico"),
        ("fin_actividades", "Fin de actividades", "Técnico"),
        ("servicio_hrs", "Servicio HRS.", "Técnico"),
        ("servicio_hh", "Servicio H/H.", "Técnico"),
        ("tipo_servicio", "Tipo de servicio", "Técnico"),
        ("tipo_reparacion", "Tipo de reparación", "Técnico"),
        ("causa_falla", "Causa de la falla", "Técnico"),
        ("actividad_realizada", "Actividad realizada", "Técnico"),
        ("refacciones_utilizadas", "Refacciones utilizadas", "Técnico"),
        ("observaciones_tecnico", "Observaciones del técnico", "Técnico"),

        ("maquina_entregada", "¿Máquina entregada?", "Entrega Supervisor"),
        ("fecha_entrega", "Fecha de entrega", "Entrega Supervisor"),
        ("observaciones_supervisor", "Observaciones del supervisor", "Entrega Supervisor")
    ]

    for nombre_campo, etiqueta, seccion in campos_base:
        existe = CampoFormato.query.filter_by(nombre_campo=nombre_campo).first()

        if not existe:
            db.session.add(CampoFormato(
                nombre_campo=nombre_campo,
                etiqueta=etiqueta,
                seccion=seccion,
                visible=True,
                es_extra=False,
                tipo_campo="texto"
            ))

    db.session.commit()


def obtener_labels(secciones):
    campos = CampoFormato.query.filter(
        CampoFormato.seccion.in_(secciones),
        CampoFormato.visible == True
    ).all()

    return {campo.nombre_campo: campo.etiqueta for campo in campos}


def obtener_campos_extra(seccion):
    return CampoFormato.query.filter_by(
        seccion=seccion,
        visible=True,
        es_extra=True
    ).all()


def guardar_valores_extra(solicitud_id, seccion):
    campos = obtener_campos_extra(seccion)

    for campo in campos:
        valor = request.form.get(f"extra_{campo.id}")

        if valor:
            existente = ValorCampoExtra.query.filter_by(
                solicitud_id=solicitud_id,
                campo_id=campo.id
            ).first()

            if existente:
                existente.valor = valor
            else:
                db.session.add(ValorCampoExtra(
                    solicitud_id=solicitud_id,
                    campo_id=campo.id,
                    valor=valor
                ))

    db.session.commit()


# =========================
# LOGIN Y SESIÓN
# =========================

@app.route("/", methods=["GET", "POST"])
def login():
    usuarios = Usuario.query.all()

    if request.method == "POST":
        rol = request.form.get("rol")
        nombre = request.form.get("nombre")
        password = request.form.get("password")

        usuario = Usuario.query.filter_by(
            rol=rol,
            nombre=nombre,
            password=password
        ).first()

        if usuario:
            session["usuario"] = usuario.nombre
            session["rol"] = usuario.rol

            if usuario.primer_login:
                return redirect(url_for("cambiar_password"))

            return redirect(url_for("menu"))

        return render_template(
            "login.html",
            usuarios=usuarios,
            error="Datos incorrectos"
        )

    return render_template("login.html", usuarios=usuarios)


@app.route("/cambiar_password", methods=["GET", "POST"])
def cambiar_password():
    if "usuario" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        nueva_password = request.form.get("password")

        usuario = Usuario.query.filter_by(
            nombre=session.get("usuario"),
            rol=session.get("rol")
        ).first()

        if usuario and nueva_password:
            usuario.password = nueva_password
            usuario.primer_login = False
            db.session.commit()

            return redirect(url_for("menu"))

    return render_template("cambiar_password.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/menu")
def menu():
    if "usuario" not in session:
        return redirect(url_for("login"))

    return render_template("menu.html")


# =========================
# SOLICITUDES
# =========================

@app.route("/nueva_orden", methods=["GET", "POST"])
def nueva_orden():
    if "usuario" not in session:
        return redirect(url_for("login"))

    if session.get("rol") not in ["admin", "supervisor"]:
        return "Acceso no autorizado"

    if request.method == "POST":
        nueva = Solicitud(
            folio=generar_folio(),
            area=request.form.get("area"),
            maquina=request.form.get("maquina"),
            fecha_solicitud=request.form.get("fecha_solicitud"),
            hora_solicitud=request.form.get("hora_solicitud"),
            solicitante=request.form.get("solicitante"),
            tecnico=request.form.get("tecnico"),
            tiempo_ciclo=request.form.get("tiempo_ciclo"),
            piezas_ciclo=request.form.get("piezas_ciclo"),
            modelo=request.form.get("modelo"),
            minutos_afectacion=request.form.get("minutos_afectacion"),
            piezas_afectadas=request.form.get("piezas_afectadas"),
            paro_produccion=request.form.get("paro_produccion"),
            fecha_paro=request.form.get("fecha_paro"),
            hora_paro=request.form.get("hora_paro"),
            fecha_arranque=request.form.get("fecha_arranque"),
            hora_arranque=request.form.get("hora_arranque"),
            descripcion=request.form.get("descripcion"),
            estado="Activa"
        )

        db.session.add(nueva)
        db.session.commit()

        guardar_valores_extra(nueva.id, "Supervisor")

        return redirect(url_for("solicitudes_activas"))

    return render_template(
        "nueva_orden.html",
        folio=generar_folio(),
        areas=Area.query.all(),
        supervisores=Supervisor.query.all(),
        tecnicos=Tecnico.query.all(),
        labels=obtener_labels(["Supervisor"]),
        campos_extra=obtener_campos_extra("Supervisor")
    )


@app.route("/solicitudes_activas")
def solicitudes_activas():
    if "usuario" not in session:
        return redirect(url_for("login"))

    if session.get("rol") not in ["admin", "tecnico"]:
        return "Acceso no autorizado"

    tecnico = request.args.get("tecnico")

    if session.get("rol") == "tecnico":
        solicitudes = Solicitud.query.filter_by(
            tecnico=session.get("usuario"),
            estado="Activa"
        ).all()
    elif tecnico and tecnico != "Todos":
        solicitudes = Solicitud.query.filter_by(
            tecnico=tecnico,
            estado="Activa"
        ).all()
    else:
        solicitudes = Solicitud.query.filter_by(estado="Activa").all()

    return render_template("solicitudes_activas.html", solicitudes=solicitudes)


@app.route("/detalle_solicitud/<int:id>", methods=["GET", "POST"])
def detalle_solicitud(id):
    if "usuario" not in session:
        return redirect(url_for("login"))

    if session.get("rol") not in ["admin", "tecnico"]:
        return "Acceso no autorizado"

    solicitud = Solicitud.query.get_or_404(id)

    if request.method == "POST":
        solicitud.participo_1 = request.form.get("participo_1")
        solicitud.participo_2 = request.form.get("participo_2")
        solicitud.participo_3 = request.form.get("participo_3")
        solicitud.inicio_actividades = request.form.get("inicio_actividades")
        solicitud.fin_actividades = request.form.get("fin_actividades")
        solicitud.servicio_hrs = request.form.get("servicio_hrs")
        solicitud.servicio_hh = request.form.get("servicio_hh")
        solicitud.tipo_servicio = request.form.get("tipo_servicio")
        solicitud.tipo_reparacion = request.form.get("tipo_reparacion")
        solicitud.causa_falla = request.form.get("causa_falla")
        solicitud.actividad_realizada = request.form.get("actividad_realizada")
        solicitud.refacciones_utilizadas = request.form.get("refacciones_utilizadas")
        solicitud.observaciones_tecnico = request.form.get("observaciones_tecnico")
        solicitud.estado = "Concluida"

        db.session.commit()

        guardar_valores_extra(solicitud.id, "Técnico")

        return redirect(url_for("solicitudes_activas"))

    return render_template(
        "detalle_solicitud.html",
        solicitud=solicitud,
        tecnicos=Tecnico.query.all(),
        labels=obtener_labels(["Supervisor", "Técnico"]),
        campos_extra=obtener_campos_extra("Técnico")
    )


@app.route("/solicitudes_concluidas")
def solicitudes_concluidas():
    if "usuario" not in session:
        return redirect(url_for("login"))

    if session.get("rol") not in ["admin", "supervisor"]:
        return "Acceso no autorizado"

    solicitudes = Solicitud.query.filter_by(estado="Concluida").all()

    return render_template(
        "solicitudes_concluidas.html",
        solicitudes=solicitudes
    )


@app.route("/entrega_supervisor/<int:id>", methods=["GET", "POST"])
def entrega_supervisor(id):
    if "usuario" not in session:
        return redirect(url_for("login"))

    if session.get("rol") not in ["admin", "supervisor"]:
        return "Acceso no autorizado"

    solicitud = Solicitud.query.get_or_404(id)

    if request.method == "POST":
        solicitud.maquina_entregada = request.form.get("maquina_entregada")
        solicitud.fecha_entrega = request.form.get("fecha_entrega")
        solicitud.observaciones_supervisor = request.form.get("observaciones_supervisor")

        if solicitud.maquina_entregada == "Sí":
            solicitud.estado = "Entregada"

        db.session.commit()

        guardar_valores_extra(solicitud.id, "Entrega Supervisor")

        return redirect(url_for("solicitudes_concluidas"))

    return render_template(
        "entrega_supervisor.html",
        solicitud=solicitud,
        labels=obtener_labels(["Supervisor", "Técnico", "Entrega Supervisor"]),
        campos_extra=obtener_campos_extra("Entrega Supervisor")
    )


# =========================
# ADMINISTRACIÓN
# =========================

@app.route("/administrar_formatos")
def administrar_formatos():
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    return render_template("administrar_formatos.html")


@app.route("/admin_control_documental", methods=["GET", "POST"])
def admin_control_documental():
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    if request.method == "POST":
        accion = request.form.get("accion")

        if accion == "agregar":
            nombre_campo = request.form.get("nombre_campo")
            etiqueta = request.form.get("etiqueta")
            seccion = request.form.get("seccion")
            tipo_campo = request.form.get("tipo_campo")

            if nombre_campo and etiqueta and seccion:
                existe = CampoFormato.query.filter_by(nombre_campo=nombre_campo).first()

                if not existe:
                    db.session.add(CampoFormato(
                        nombre_campo=nombre_campo,
                        etiqueta=etiqueta,
                        seccion=seccion,
                        tipo_campo=tipo_campo,
                        visible=True,
                        es_extra=True
                    ))

                    db.session.commit()

        if accion == "guardar":
            for campo in CampoFormato.query.all():
                etiqueta = request.form.get(f"etiqueta_{campo.id}")

                if etiqueta:
                    campo.etiqueta = etiqueta

                campo.visible = request.form.get(f"visible_{campo.id}") == "on"

            db.session.commit()

        return redirect(url_for("admin_control_documental"))

    return render_template(
        "admin_control_documental.html",
        campos_supervisor=CampoFormato.query.filter_by(seccion="Supervisor").all(),
        campos_tecnico=CampoFormato.query.filter_by(seccion="Técnico").all(),
        campos_entrega=CampoFormato.query.filter_by(seccion="Entrega Supervisor").all()
    )


@app.route("/admin_areas", methods=["GET", "POST"])
def admin_areas():
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    if request.method == "POST":
        nombre = request.form.get("nombre")

        if nombre:
            db.session.add(Area(nombre=nombre))
            db.session.commit()

        return redirect(url_for("admin_areas"))

    return render_template("admin_areas.html", areas=Area.query.all())


@app.route("/editar_area/<int:id>", methods=["POST"])
def editar_area(id):
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    area = Area.query.get_or_404(id)
    area.nombre = request.form.get("nombre")
    db.session.commit()

    return redirect(url_for("admin_areas"))


@app.route("/eliminar_area/<int:id>")
def eliminar_area(id):
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    area = Area.query.get_or_404(id)
    db.session.delete(area)
    db.session.commit()

    return redirect(url_for("admin_areas"))


@app.route("/admin_supervisores", methods=["GET", "POST"])
def admin_supervisores():
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    if request.method == "POST":
        nombre = request.form.get("nombre")

        if nombre:
            db.session.add(Supervisor(nombre=nombre))
            db.session.commit()

        return redirect(url_for("admin_supervisores"))

    return render_template("admin_supervisores.html", supervisores=Supervisor.query.all())


@app.route("/editar_supervisor/<int:id>", methods=["POST"])
def editar_supervisor(id):
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    supervisor = Supervisor.query.get_or_404(id)
    supervisor.nombre = request.form.get("nombre")
    db.session.commit()

    return redirect(url_for("admin_supervisores"))


@app.route("/eliminar_supervisor/<int:id>")
def eliminar_supervisor(id):
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    supervisor = Supervisor.query.get_or_404(id)
    db.session.delete(supervisor)
    db.session.commit()

    return redirect(url_for("admin_supervisores"))


@app.route("/admin_tecnicos", methods=["GET", "POST"])
def admin_tecnicos():
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    if request.method == "POST":
        nombre = request.form.get("nombre")

        if nombre:
            db.session.add(Tecnico(nombre=nombre))
            db.session.commit()

        return redirect(url_for("admin_tecnicos"))

    return render_template("admin_tecnicos.html", tecnicos=Tecnico.query.all())


@app.route("/editar_tecnico/<int:id>", methods=["POST"])
def editar_tecnico(id):
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    tecnico = Tecnico.query.get_or_404(id)
    tecnico.nombre = request.form.get("nombre")
    db.session.commit()

    return redirect(url_for("admin_tecnicos"))


@app.route("/eliminar_tecnico/<int:id>")
def eliminar_tecnico(id):
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    tecnico = Tecnico.query.get_or_404(id)
    db.session.delete(tecnico)
    db.session.commit()

    return redirect(url_for("admin_tecnicos"))


@app.route("/admin_usuarios", methods=["GET", "POST"])
def admin_usuarios():
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    if request.method == "POST":
        nombre = request.form.get("nombre")
        rol = request.form.get("rol")
        password = request.form.get("password")

        if nombre and rol and password:
            nuevo = Usuario(
                nombre=nombre,
                rol=rol,
                password=password,
                primer_login=True
            )

            db.session.add(nuevo)
            db.session.commit()

        return redirect(url_for("admin_usuarios"))

    usuarios = Usuario.query.all()

    return render_template("admin_usuarios.html", usuarios=usuarios)


@app.route("/editar_usuario/<int:id>", methods=["POST"])
def editar_usuario(id):
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    usuario = Usuario.query.get_or_404(id)

    usuario.nombre = request.form.get("nombre")
    usuario.rol = request.form.get("rol")

    nueva_password = request.form.get("password")

    if nueva_password:
        usuario.password = nueva_password
        usuario.primer_login = True

    db.session.commit()

    return redirect(url_for("admin_usuarios"))


@app.route("/eliminar_usuario/<int:id>")
def eliminar_usuario(id):
    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    usuario = Usuario.query.get_or_404(id)
    db.session.delete(usuario)
    db.session.commit()

    return redirect(url_for("admin_usuarios"))


# =========================
# EXPORTAR EXCEL
# =========================

@app.route("/exportar_excel")
def exportar_excel():
    if "usuario" not in session:
        return redirect(url_for("login"))

    if session.get("rol") != "admin":
        return "Acceso no autorizado"

    solicitudes = Solicitud.query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Mantenimiento"

    encabezados = [
        "Folio", "Área", "Máquina", "Solicitante", "Técnico Asignado",
        "Fecha Solicitud", "Hora Solicitud", "Descripción de Falla",
        "Causa de Falla", "Actividad Realizada", "Refacciones Utilizadas",
        "Inicio Actividades", "Fin Actividades", "Servicio HRS", "Servicio H/H",
        "Tipo Servicio", "Tipo Reparación", "Estado", "Máquina Entregada",
        "Fecha Entrega", "Observaciones Supervisor"
    ]

    ws.append(encabezados)

    color_encabezado = "1F4E78"
    color_fila = "D9EAF7"
    blanco = "FFFFFF"

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = PatternFill(start_color=color_encabezado, end_color=color_encabezado, fill_type="solid")
        cell.font = Font(color=blanco, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde

    for s in solicitudes:
        ws.append([
            s.folio, s.area, s.maquina, s.solicitante, s.tecnico,
            s.fecha_solicitud, s.hora_solicitud, s.descripcion,
            s.causa_falla, s.actividad_realizada, s.refacciones_utilizadas,
            s.inicio_actividades, s.fin_actividades, s.servicio_hrs,
            s.servicio_hh, s.tipo_servicio, s.tipo_reparacion, s.estado,
            s.maquina_entregada, s.fecha_entrega, s.observaciones_supervisor
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill(start_color=color_fila, end_color=color_fila, fill_type="solid")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 4, 35)

    ws.row_dimensions[1].height = 35

    archivo = "reporte_mantenimiento_Correctivo.xlsx"
    wb.save(archivo)

    return send_file(archivo, as_attachment=True)


# =========================
# INICIO DE APP
# =========================

with app.app_context():
    db.create_all()
    crear_datos_iniciales()
    crear_campos_formato_iniciales()


if __name__ == "__main__":
    app.run(debug=True)
